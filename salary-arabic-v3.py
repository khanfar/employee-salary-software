import os
import tkinter as tk
from tkinter import messagebox, simpledialog
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
from tkinter import ttk
import csv
import openpyxl
from openpyxl.styles import Alignment

data_loaded = False  # Initialize the data_loaded variable
employees = []
search_results = []  # Declare search_results as a global variable


def load_employee_data(file_path):
    employee_data = {}
    with open(file_path, encoding='utf-8') as file:
        for line in file:
            name, salary, date = line.strip().split(',')
            date = datetime.strptime(date, '%Y-%m-%d')
            salary = float(salary)

            if name not in employee_data:
                employee_data[name] = []

            employee_data[name].append((date, salary))

    return employee_data


def load_your_data(file_path):
    entries = []
    with open(file_path, encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            if len(row) > 0:
                entries.append(row)
    return entries


def calculate_monthly_income(entries, month, year):
    income_data = {}
    for entry in entries:
        try:
            vehicle_plate, vehicle_type, entry_date, company, driver, phone, mechanic, note1, note2, final_report = entry
            entry_date = datetime.strptime(entry_date, '%d.%m.%Y')

            if entry_date.month == month and entry_date.year == year:
                mechanic = mechanic.strip()
                if 'شيكل' in final_report:
                    amount_str = final_report.split('شيكل')[0].strip().split()[-1]
                    if 'شغل' in final_report:
                        amount_str = final_report.split('شغل')[1].split('شيكل')[0].strip()
                    amount = int(amount_str)
                    if mechanic:
                        if mechanic not in income_data:
                            income_data[mechanic] = 0
                        income_data[mechanic] += amount
        except Exception as e:
            continue
    return income_data


def calculate_range_income(entries, start_date, end_date):
    income_data = {}
    for entry in entries:
        try:
            vehicle_plate, vehicle_type, entry_date, company, driver, phone, mechanic, note1, note2, final_report = entry
            entry_date = datetime.strptime(entry_date, '%d.%m.%Y')

            if start_date <= entry_date <= end_date:
                mechanic = mechanic.strip()
                if 'شيكل' in final_report:
                    amount_str = final_report.split('شيكل')[0].strip().split()[-1]
                    if 'شغل' in final_report:
                        amount_str = final_report.split('شغل')[1].split('شيكل')[0].strip()
                    amount = int(amount_str)
                    if mechanic:
                        if mechanic not in income_data:
                            income_data[mechanic] = 0
                        income_data[mechanic] += amount
        except Exception as e:
            continue
    return income_data


def generate_report(employee_data, income_data, month, year, output_file_txt, output_file_xlsx):
    with open(output_file_txt, 'w', encoding='utf-8') as file:
        # Write month and year at the top
        file.write(f"Report for {year}-{month:02d}\n")
        file.write("="*40 + "\n")

        for employee, salary_entries in employee_data.items():
            total_salary = sum(salary for date, salary in salary_entries if date.month == month and date.year == year)
            total_income = income_data.get(employee, 0)
            profit = total_income - total_salary
            file.write(f"{employee}, Total Salary: {total_salary}, Total Income: {total_income}, Profit: {profit}\n")

    # Excel report generation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {year}-{month:02d}"

    # Write report title
    ws['A1'] = f"Report for {year}-{month:02d}"
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Writing headers in Arabic
    ws.append(["موظف", "الراتب الإجمالي", "الدخل الإجمالي", "الأرباح"])

    # Writing data
    for employee, salary_entries in employee_data.items():
        total_salary = sum(salary for date, salary in salary_entries if date.month == month and date.year == year)
        total_income = income_data.get(employee, 0)
        profit = total_income - total_salary
        ws.append([employee, total_salary, total_income, profit])

    wb.save(output_file_xlsx)


def generate_report_range(employee_data, income_data, start_date, end_date, output_file_txt, output_file_xlsx):
    with open(output_file_txt, 'w', encoding='utf-8') as file:
        # Write date range at the top
        file.write(f"Report for {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}\n")
        file.write("=" * 40 + "\n")

        for employee, salary_entries in employee_data.items():
            total_salary = sum(salary for date, salary in salary_entries if start_date <= date <= end_date)
            total_income = income_data.get(employee, 0)
            profit = total_income - total_salary
            file.write(
                f"{employee}, Total Salary: {total_salary}, Total Income: {total_income}, Profit: {profit}\n")

    # Excel report generation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"

    # Write report title
    ws['A1'] = f"Report for {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Writing headers in Arabic
    ws.append(["موظف", "الراتب الإجمالي", "الدخل الإجمالي", "الأرباح"])

    # Writing data
    for employee, salary_entries in employee_data.items():
        total_salary = sum(salary for date, salary in salary_entries if start_date <= date <= end_date)
        total_income = income_data.get(employee, 0)
        profit = total_income - total_salary
        ws.append([employee, total_salary, total_income, profit])

    wb.save(output_file_xlsx)


def main():
    employee_file = 'employee_data.txt'
    your_data_file = 'your_data.csv'
    output_dir = 'output'

    # Ensure the output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    choice = input("Do you want to specify a single month (1) or a range of dates (2)? Enter 1 or 2: ")

    if choice == '1':
        month = int(input("Enter the month (1-12): "))
        year = int(input("Enter the year (e.g., 2024): "))
        output_file_txt = os.path.join(output_dir, f'report_{year}_{month:02d}.txt')
        output_file_xlsx = os.path.join(output_dir, f'report_{year}_{month:02d}.xlsx')

        employee_data = load_employee_data(employee_file)
        your_data_entries = load_your_data(your_data_file)
        income_data = calculate_monthly_income(your_data_entries, month, year)
        generate_report(employee_data, income_data, month, year, output_file_txt, output_file_xlsx)

    elif choice == '2':
        start_date_str = input("Enter the start date (YYYY-MM-DD): ")
        end_date_str = input("Enter the end date (YYYY-MM-DD): ")

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        output_file_txt = os.path.join(output_dir, f'report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.txt')
        output_file_xlsx = os.path.join(output_dir, f'report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx')

        employee_data = load_employee_data(employee_file)
        your_data_entries = load_your_data(your_data_file)
        income_data = calculate_range_income(your_data_entries, start_date, end_date)
        generate_report_range(employee_data, income_data, start_date, end_date, output_file_txt, output_file_xlsx)

    else:
        print("Invalid choice. Please run the program again and enter either 1 or 2.")


# GUI Functionality

# Function to add a new employee to the database
def add_employee():
    name = name_entry.get()
    try:
        salary = float(salary_entry.get())
    except ValueError:
        messagebox.showerror("خطأ", "الرجاء إدخال راتب صحيح.")
        return

    current_time = datetime.now().strftime("%Y-%m-%d")
    employees.append({'name': name, 'salary': salary, 'date': current_time})
    update_display()
    # Clear entry fields
    name_entry.delete(0, tk.END)
    salary_entry.delete(0, tk.END)

    # Automatically save data after adding an employee
    save_data()


# Function to delete an employee from the database
def delete_employee():
    index = text.curselection()
    if index:
        confirm = messagebox.askyesno("تأكيد", "هل أنت متأكد أنك تريد حذف هذا الموظف؟")
        if confirm:
            del employees[index[0]]
            update_display()
            # Automatically save data after deleting an employee
            save_data()
    else:
        messagebox.showerror("خطأ", "الرجاء تحديد موظف لحذفه.")


# Function to update the salary of an existing employee
def update_salary():
    index = text.curselection()
    if index:
        try:
            new_salary = float(new_salary_entry.get())
            employees[index[0]]['salary'] = new_salary
            update_display()
            # Automatically save data after update a salary
            save_data()
        except ValueError:
            messagebox.showerror("خطأ", "الرجاء إدخال راتب صحيح.")
    else:
        messagebox.showerror("خطأ", "الرجاء تحديد موظف لتحديث راتبه.")


# Function to save employee data to a file
def save_data():
    global data_loaded  # Declare data_loaded as global
    if not os.path.exists("employee_data.txt") or (os.path.exists("employee_data.txt") and data_loaded):
        with open("employee_data.txt", "w") as file:
            for employee in employees:
                file.write(f"{employee['name']},{employee['salary']},{employee['date']}\n")
        messagebox.showinfo("نجاح", "تم حفظ بيانات الموظفين بنجاح.")
    else:
        messagebox.showinfo("معلومات", "لم يتم تحميل بيانات الموظفين. لا يمكن الحفظ.")


# Function to load employee data from a file
def load_data():
    global data_loaded  # Declare data_loaded as global
    if not data_loaded:  # Check if data is not already loaded
        try:
            with open("employee_data.txt", "r") as file:
                lines = file.readlines()
                for line in lines:
                    data = line.strip().split(',')
                    if len(data) == 3:  # Ensure all required fields are present
                        name, salary, date = data
                        employees.append({'name': name, 'salary': float(salary), 'date': date})
                    else:
                        messagebox.showwarning("تحذير", "بعض البيانات في الملف غير صالحة.")
            update_display()
            messagebox.showinfo("نجاح", "تم تحميل بيانات الموظفين بنجاح.")
            data_loaded = True  # Set the flag to True after loading data
        except FileNotFoundError:
            messagebox.showerror("خطأ", "لم يتم العثور على بيانات الموظفين.")
            data_loaded = False  # Set the flag to False if no data file is found
    else:
        messagebox.showinfo("معلومات", "تم تحميل بيانات الموظفين بالفعل.")


# Function to sort employees by name or salary
def sort_employees(sort_key):
    employees.sort(key=lambda x: x[sort_key])
    update_display()


# Function to search for employees by name or date
def search_employee():
    global search_results  # Declare search_results as global
    query = search_entry.get().lower()

    # Search by name
    name_results = [employee for employee in employees if query in employee['name'].lower()]
    date_results = []

    # If not found in name, check if it's a valid date
    if not name_results:
        try:
            query_date = datetime.strptime(query, '%d-%m-%Y').strftime('%Y-%m-%d')
            date_results = [employee for employee in employees if query_date in employee['date']]
        except ValueError:
            try:
                query_date = datetime.strptime(query, '%d.%m.%Y').strftime('%Y-%m-%d')
                date_results = [employee for employee in employees if query_date in employee['date']]
            except ValueError:
                pass

    search_results = name_results + date_results

    if search_results:
        text.delete(0, tk.END)
        for employee in search_results:
            text.insert(tk.END, f"الاسم: {employee['name']}, الراتب: {employee['salary']}, التاريخ: {employee.get('date', 'غير متاح')}\n")
    else:
        messagebox.showinfo("البحث", "لا يوجد موظفين مطابقين.")


# Function to export filtered employee data to Excel and text file in a folder named "output"
def export_data():
    search_query = search_entry.get().lower()
    search_results = []

    # Search for matching employees
    if search_query:
        # Search by name
        name_results = [employee for employee in employees if search_query in employee['name'].lower()]
        date_results = []

        # If not found in name, check if it's a valid date
        if not name_results:
            try:
                query_date = datetime.strptime(search_query, '%d-%m-%Y').strftime('%Y-%m-%d')
                date_results = [employee for employee in employees if query_date in employee['date']]
            except ValueError:
                try:
                    query_date = datetime.strptime(search_query, '%d.%m.%Y').strftime('%Y-%m-%d')
                    date_results = [employee for employee in employees if query_date in employee['date']]
                except ValueError:
                    pass

        search_results = name_results + date_results
    else:
        search_results = employees

    if search_results:
        # Create the output folder if it doesn't exist
        if not os.path.exists("output"):
            os.makedirs("output")

        # Export data to Excel file
        excel_file_path = os.path.join("output", "employee_data.xlsx")
        df = pd.DataFrame(search_results)
        df.to_excel(excel_file_path, index=False)

        # Export data to text file
        txt_file_path = os.path.join("output", "employee_data.txt")
        with open(txt_file_path, "w") as file:
            for employee in search_results:
                # Format salary without decimal point and trailing zeros
                formatted_salary = f"{employee['salary']:.0f}"
                file.write(f"{employee['name']},{formatted_salary},{employee['date']}\n")

        messagebox.showinfo("التصدير", "تم تصدير بيانات الموظفين المصفاة إلى ملف Excel وملف نصي في مجلد 'output' بنجاح.")
    else:
        messagebox.showinfo("التصدير", "لا توجد بيانات موظف مطابقة متاحة للتصدير.")


# Function to export filtered employee data to a text file by a specific date
def export_data_by_date():
    today_date = datetime.now().strftime('%Y-%m-%d')
    user_choice = messagebox.askyesno("تصدير حسب التاريخ", f"هل ترغب في تصدير البيانات لتاريخ اليوم ({today_date})؟")

    if user_choice:
        search_results = [employee for employee in employees if employee['date'] == today_date]
        specific_date = today_date  # Set specific_date to today_date
    else:
        specific_date = simpledialog.askstring("تصدير حسب التاريخ", "أدخل التاريخ (YYYY-MM-DD):")
        if specific_date is None:  # If user cancels the operation, return early
            return

        search_results = [employee for employee in employees if employee['date'] == specific_date]

    if search_results:
        # Calculate total salary
        total_salary = sum(employee['salary'] for employee in search_results)

        # Create the output folder if it doesn't exist
        if not os.path.exists("output"):
            os.makedirs("output")

        # Export data to text file
        txt_file_path = os.path.join("output", f"employee_data_{specific_date}.txt")
        with open(txt_file_path, "w") as file:
            # Header
            file.write(" شركه ابناء عرفات \n")
            file.write(" رواتب الموظفين \n")
            file.write(f" {specific_date} \n\n")

            # Employee data
            for employee in search_results:
                formatted_salary = f"{employee['salary']} شيكل"
                file.write(f"{employee['name'].ljust(10)}{formatted_salary.rjust(20)}\n")

            # Separate line
            file.write("\n*************************\n")

            # Total amount
            total_line = f"المجموع الكلي: {total_salary} شيكل"
            file.write(total_line.center(40) + "\n")

            # Separate line
            file.write("*************************\n")

        messagebox.showinfo("تصدير حسب التاريخ", "تم تصدير بيانات الموظفين للتاريخ المحدد إلى ملف نصي في مجلد 'output' بنجاح.")
    else:
        messagebox.showinfo("تصدير حسب التاريخ", "لا توجد بيانات موظف متاحة للتصدير للتاريخ المحدد.")


# Function to collect total salary of search results and display it in a popup window
def collect_results():
    if search_results:
        total_salary = sum(employee['salary'] for employee in search_results)
        messagebox.showinfo("نتائج البحث", f"إجمالي رواتب الموظفين: {total_salary} شيكل")
    else:
        messagebox.showinfo("نتائج البحث", "لا توجد نتائج بحث حالياً.")


# Function to visualize employee data
def visualize_data():
    global search_results  # Declare search_results as global
    if employees:
        if search_results:
             df = pd.DataFrame(search_results)
             df.plot(kind='bar', x='name', y='salary', rot=45)
             plt.xlabel('الموظف')
             plt.ylabel('الراتب')
             plt.title('توزيع رواتب الموظفين')
             plt.tight_layout()
             plt.show()
        else:
            messagebox.showinfo("تصور البيانات", "لا توجد نتائج بحث للتصور.")
    else:
        messagebox.showinfo("تصور البيانات", "لا توجد بيانات موظف متاحة للتصور.")


# Function to update the display with current employee data
def update_display():
    text.delete(0, tk.END)
    for employee in employees:
        text.insert(tk.END, f"الاسم: {employee['name']}, الراتب: {employee['salary']}, التاريخ: {employee['date']}\n")


# Function to export filtered employee data to a text file by a specific month and year
def export_data_by_month():
    month_year = simpledialog.askstring("تصدير حسب الشهر", "أدخل الشهر والسنة (MM/YYYY):")
    if month_year is None:  # If user cancels the operation, return early
        return

    try:
        month, year = map(int, month_year.split('/'))
        if not (1 <= month <= 12 and year >= 1900):
            raise ValueError
    except ValueError:
        messagebox.showerror("خطأ", "يرجى إدخال شهر وسنة صالحين (MM/YYYY).")
        return

    # Filter employees by month and year
    search_results = [employee for employee in employees if datetime.strptime(employee['date'], '%Y-%m-%d').month == month and datetime.strptime(employee['date'], '%Y-%m-%d').year == year]

    if search_results:
        # Calculate total salary
        total_salary = sum(employee['salary'] for employee in search_results)

        # Create the output folder if it doesn't exist
        if not os.path.exists("output"):
            os.makedirs("output")

        # Export data to text file
        txt_file_path = os.path.join("output", f"employee_data_{month:02d}_{year}.txt")
        with open(txt_file_path, "w") as file:
            # Header
            file.write(" شركه ابناء عرفات \n")
            file.write(" رواتب الموظفين \n")
            file.write(f" {month:02d}/{year} \n\n")

            # Employee data
            for employee in search_results:
                formatted_salary = f"{employee['salary']} شيكل"
                file.write(f"{employee['name'].ljust(10)}{formatted_salary.rjust(20)}\n")

            # Separate line
            file.write("\n*************************\n")

            # Total amount
            total_line = f"المجموع الكلي: {total_salary} شيكل"
            file.write(total_line.center(40) + "\n")

            # Separate line
            file.write("*************************\n")

        messagebox.showinfo("تصدير حسب الشهر", f"بيانات الموظفين لشهر {month:02d}/{year} تم تصديرها إلى ملف نصي بنجاح في مجلد 'output'.")
    else:
        messagebox.showinfo("تصدير حسب الشهر", f"لا تتوفر بيانات الموظفين لشهر {month:02d}/{year}.")


# Function to export filtered employee data to a text file by a specific date range
def export_data_by_date_range():
    start_date = simpledialog.askstring("تصدير من تاريخ إلى تاريخ", "أدخل تاريخ البداية (YYYY-MM-DD):")
    if start_date is None:  # If user cancels the operation, return early
        return

    end_date = simpledialog.askstring("تصدير من تاريخ إلى تاريخ", "أدخل تاريخ النهاية (YYYY-MM-DD):")
    if end_date is None:  # If user cancels the operation, return early
        return

    try:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        if start > end:
            raise ValueError
    except ValueError:
        messagebox.showerror("خطأ", "يرجى إدخال تواريخ صالحة.")
        return

    # Filter employees by date range
    search_results = [employee for employee in employees if start <= datetime.strptime(employee['date'], '%Y-%m-%d') <= end]

    if search_results:
        # Create the output folder if it doesn't exist
        if not os.path.exists("output"):
            os.makedirs("output")

        # Export data to text file
        txt_file_path = os.path.join("output", f"employee_data_{start_date}_to_{end_date}.txt")
        with open(txt_file_path, "w") as file:
            # Header
            file.write(" شركه ابناء عرفات \n")
            file.write(" رواتب الموظفين \n")
            file.write(f" من {start_date} إلى {end_date} \n\n")

            # Employee data
            for employee in search_results:
                formatted_salary = f"{employee['salary']} شيكل"
                file.write(f"{employee['name'].ljust(10)}{formatted_salary.rjust(20)}\n")

            # Separate line
            file.write("\n*************************\n")

            # Total amount
            total_salary = sum(employee['salary'] for employee in search_results)
            total_line = f"المجموع الكلي: {total_salary} شيكل"
            file.write(total_line.center(40) + "\n")

            # Separate line
            file.write("*************************\n")

        messagebox.showinfo("تصدير من تاريخ إلى تاريخ", f"تم تصدير بيانات الموظفين من {start_date} إلى {end_date} إلى ملف نصي بنجاح في مجلد 'output'.")
    else:
        messagebox.showinfo("تصدير من تاريخ إلى تاريخ", f"لا توجد بيانات موظف متاحة للتصدير للفترة من {start_date} إلى {end_date}.")


# Function to calculate profit
def calculate_profit():
    employee_file = 'employee_data.txt'
    your_data_file = 'your_data.csv'
    output_dir = 'output'

    # Ensure the output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    choice = simpledialog.askstring("تحديد الفترة", "هل تريد تحديد شهر واحد (1) أم نطاق تواريخ (2)؟ أدخل 1 أو 2:")

    if choice == '1':
        month = simpledialog.askinteger("الشهر", "أدخل الشهر (1-12):")
        year = simpledialog.askinteger("السنة", "أدخل السنة (مثلًا، 2024):")
        output_file_txt = os.path.join(output_dir, f'report_{year}_{month:02d}.txt')
        output_file_xlsx = os.path.join(output_dir, f'report_{year}_{month:02d}.xlsx')

        employee_data = load_employee_data(employee_file)
        your_data_entries = load_your_data(your_data_file)
        income_data = calculate_monthly_income(your_data_entries, month, year)
        generate_report(employee_data, income_data, month, year, output_file_txt, output_file_xlsx)
        messagebox.showinfo("حساب الأرباح", f"تم حساب الأرباح للشهر {month:02d}/{year} وحفظ التقرير في مجلد 'output'.")

    elif choice == '2':
        start_date_str = simpledialog.askstring("تاريخ البداية", "أدخل تاريخ البداية (YYYY-MM-DD):")
        end_date_str = simpledialog.askstring("تاريخ النهاية", "أدخل تاريخ النهاية (YYYY-MM-DD):")

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        output_file_txt = os.path.join(output_dir, f'report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.txt')
        output_file_xlsx = os.path.join(output_dir, f'report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx')

        employee_data = load_employee_data(employee_file)
        your_data_entries = load_your_data(your_data_file)
        income_data = calculate_range_income(your_data_entries, start_date, end_date)
        generate_report_range(employee_data, income_data, start_date, end_date, output_file_txt, output_file_xlsx)
        messagebox.showinfo("حساب الأرباح", f"تم حساب الأرباح من {start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')} وحفظ التقرير في مجلد 'output'.")

    else:
        messagebox.showerror("خطأ", "اختيار غير صالح. يرجى إعادة المحاولة وإدخال إما 1 أو 2.")


# GUI Setup
root = tk.Tk()
root.title("إدارة رواتب الموظفين")

# Label and entry for adding new employees
name_label = tk.Label(root, text="الاسم:")
name_label.grid(row=0, column=0, padx=10, pady=5, sticky=tk.E)

name_entry = tk.Entry(root)
name_entry.grid(row=0, column=1, padx=10, pady=5)

salary_label = tk.Label(root, text="الراتب:")
salary_label.grid(row=1, column=0, padx=10, pady=5, sticky=tk.E)

salary_entry = tk.Entry(root)
salary_entry.grid(row=1, column=1, padx=10, pady=5)

add_button = tk.Button(root, text="إضافة موظف", command=add_employee)
add_button.grid(row=2, column=0, columnspan=2, padx=10, pady=5)

# Text widget to display employee data
text = tk.Listbox(root, width=60, height=15)
text.grid(row=3, column=0, columnspan=2, padx=10, pady=5, sticky=tk.W + tk.E + tk.N + tk.S)

# Button to delete selected employee
delete_button = tk.Button(root, text="حذف موظف", command=delete_employee)
delete_button.grid(row=4, column=0, columnspan=2, padx=10, pady=5)

# Label and entry for updating salary
new_salary_label = tk.Label(root, text="الراتب الجديد:")
new_salary_label.grid(row=5, column=0, padx=10, pady=5, sticky=tk.E)

new_salary_entry = tk.Entry(root)
new_salary_entry.grid(row=5, column=1, padx=10, pady=5)

update_salary_button = tk.Button(root, text="تحديث الراتب", command=update_salary)
update_salary_button.grid(row=6, column=0, columnspan=2, padx=10, pady=5)

# Button to save and load data
save_button = tk.Button(root, text="حفظ البيانات", command=save_data)
save_button.grid(row=7, column=0, padx=10, pady=5)

load_button = tk.Button(root, text="تحميل البيانات", command=load_data)
load_button.grid(row=7, column=1, padx=10, pady=5)

# Buttons for sorting employees
sort_name_button = tk.Button(root, text="ترتيب حسب الاسم", command=lambda: sort_employees('name'))
sort_name_button.grid(row=8, column=0, columnspan=2, padx=10, pady=5)

sort_salary_button = tk.Button(root, text="ترتيب حسب الراتب", command=lambda: sort_employees('salary'))
sort_salary_button.grid(row=9, column=0, columnspan=2, padx=10, pady=5)

# Entry and button for searching employees
search_entry = tk.Entry(root)
search_entry.grid(row=10, column=0, padx=10, pady=5)

search_button = tk.Button(root, text="بحث", command=search_employee)
search_button.grid(row=10, column=1, padx=10, pady=5)

# Button for data visualization
visualize_button = tk.Button(root, text="تصور البيانات", command=visualize_data)
visualize_button.grid(row=11, column=0, columnspan=2, padx=10, pady=5)

# Button for exporting data to Excel and text file
export_excel_button = tk.Button(root, text="تصدير إلى Excel وملف نصي", command=export_data)
export_excel_button.grid(row=12, column=0, columnspan=2, padx=10, pady=5)

# Button for exporting data by specific date
export_by_date_button = tk.Button(root, text="تصدير حسب التاريخ", command=export_data_by_date)
export_by_date_button.grid(row=15, column=0, padx=5, pady=5)

# Button for exporting data by specific month and year
export_by_month_button = tk.Button(root, text="تصدير حسب الشهر", command=export_data_by_month)
export_by_month_button.grid(row=15, column=1, padx=5, pady=5)

# Button for exporting data from a specific start date to an end date
export_by_date_range_button = tk.Button(root, text="تصدير من تاريخ إلى تاريخ", command=export_data_by_date_range)
export_by_date_range_button.grid(row=15, column=2, padx=5, pady=5)

# Button to collect total search results
collect_results_button = tk.Button(root, text="جمع النتائج", command=collect_results)
collect_results_button.grid(row=11, column=1, padx=5, pady=5)

# Button to calculate profit
profit_button = tk.Button(root, text="حساب الأرباح", command=calculate_profit)
profit_button.grid(row=16, column=0, columnspan=2, padx=10, pady=5)

# Center the window on the screen
root.update_idletasks()
root_width = root.winfo_width()
root_height = root.winfo_height()
x_offset = (root.winfo_screenwidth() - root_width) // 2
y_offset = (root.winfo_screenheight() - root_height) // 2
root.geometry(f"{root_width}x{root_height}+{x_offset}+{y_offset}")

root.mainloop()
